import psutil
import datetime
from openpyxl import load_workbook

EXCEL_PATH = r"C:\Users\Adejo\iCloudDrive\IT Learning Log\Updated IT Practice Log.xlsx"

def _log_to_excel(data: dict):
    """Log system health metrics to the 'System Health' sheet in the Excel workbook."""
    wb = load_workbook(EXCEL_PATH)
    if "System Health" in wb.sheetnames:
        ws = wb["System Health"]
    else:
        ws = wb.create_sheet("System Health")
        headers = ["Timestamp", "CPU %", "RAM %", "Disk % Free", "Top Processes"]
        ws.append(headers)
    ws.append([
        datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        data.get("cpu"),
        data.get("ram"),
        data.get("disk_free"),
        ", ".join(data.get("top_procs", []))
    ])
    wb.save(EXCEL_PATH)

def _top_heavy_processes(threshold: float = 85.0, top_n: int = 3):
    """Return a list of strings describing top processes by CPU or RAM if usage exceeds threshold."""
    heavy = []
    # Build list of (name, cpu%, ram%)
    for p in psutil.process_iter(['pid', 'name', 'cpu_percent', 'memory_percent']):
        try:
            cpu = p.info['cpu_percent']
            ram = p.info['memory_percent']
            if cpu > threshold or ram > threshold:
                heavy.append((p.info['name'], cpu, ram))
        except (psutil.NoSuchProcess, psutil.AccessDenied):
            continue
    # Sort by max of cpu or ram usage descending
    heavy.sort(key=lambda x: max(x[1], x[2]), reverse=True)
    return [f"{name} (CPU {cpu:.1f}%, RAM {ram:.1f}%)" for name, cpu, ram in heavy[:top_n]]

def check_system_health():
    """Gather CPU, RAM, and disk usage. Log to Excel and return a human‑readable summary."""
    cpu = psutil.cpu_percent(interval=1)
    ram = psutil.virtual_memory().percent
    disk = psutil.disk_usage('/').percent
    disk_free = 100 - disk
    top = []
    if cpu > 85.0 or ram > 85.0:
        top = _top_heavy_processes()
    data = {
        "cpu": cpu,
        "ram": ram,
        "disk_free": disk_free,
        "top_procs": top,
    }
    _log_to_excel(data)
    summary = f"CPU {cpu:.1f}% | RAM {ram:.1f}% | Disk free {disk_free:.1f}%"
    if top:
        summary += " | Heavy processes: " + ", ".join(top)
    return summary
