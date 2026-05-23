import openpyxl
from datetime import datetime

EXCEL_PATH = r"C:\Users\Adejo\iCloudDrive\IT Learning Log\Updated IT Practice Log.xlsx"

def log_session(topic: str, description: str, completed: bool, duration: str) -> str:
    """Log a lab session to the Excel workbook.

    Args:
        topic: The session topic.
        description: What was done.
        completed: Whether the session was completed.
        duration: Duration in minutes.
    Returns:
        Confirmation message string.
    """
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    now = datetime.now()
    date = now.strftime("%d/%m/%Y")
    day = now.strftime("%A")
    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=1, value=date)
    ws.cell(row=next_row, column=2, value=day)
    ws.cell(row=next_row, column=3, value=topic)
    ws.cell(row=next_row, column=4, value=description)
    ws.cell(row=next_row, column=5, value=completed)
    ws.cell(row=next_row, column=6, value=duration)
    wb.save(EXCEL_PATH)
    return f"Session logged successfully — {date}, {topic}, duration {duration} minutes"

if __name__ == "__main__":
    # Example interactive usage (not used as a tool)
    topic = input("What topic did you cover today? ")
    description = input("What did you do? ")
    completed_input = input("Was it completed? (yes/no) ").strip().lower()
    completed = True if completed_input == "yes" else False
    duration = input("How long (in minutes) was the session? ")
    print(log_session(topic, description, completed, duration))