import subprocess
import socket
import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

EXCEL_PATH = r"C:\Users\Adejo\iCloudDrive\IT Learning Log\Updated IT Practice Log.xlsx"

def _log_to_excel(data: dict):
    """Log network check results to the 'Network Logs' sheet in the Excel workbook."""
    wb = load_workbook(EXCEL_PATH)
    if "Network Logs" in wb.sheetnames:
        ws = wb["Network Logs"]
    else:
        ws = wb.create_sheet("Network Logs")
        headers = ["Timestamp", "Target", "Ping (ms)", "Packet Loss (%)", "DNS", "Traceroute", "Open Ports"]
        ws.append(headers)
    ws.append([
        datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        data.get("target"),
        data.get("ping_ms"),
        data.get("packet_loss"),
        ", ".join(data.get("dns", [])),
        data.get("tracert", ""),
        ", ".join(map(str, data.get("open_ports", [])))
    ])
    wb.save(EXCEL_PATH)

def ping_target(target: str, count: int = 4) -> dict:
    """Ping a target and return latency (ms) and packet loss (%)."""
    try:
        result = subprocess.run(["ping", target, "-n", str(count)], capture_output=True, text=True, timeout=10)
        output = result.stdout
        # Extract average time and loss percentage
        loss_line = [l for l in output.splitlines() if "Lost =" in l]
        loss = 0
        if loss_line:
            parts = loss_line[0].split(",")
            for p in parts:
                if "Lost =" in p:
                    loss = int(p.split("=")[1].strip().split("(")[0])
        # Average time line
        avg_time = None
        for l in output.splitlines():
            if "Average =" in l:
                avg_time = int(l.split("=")[1].replace("ms", "").strip())
                break
        return {"ping_ms": avg_time, "packet_loss": loss}
    except Exception as e:
        return {"ping_ms": None, "packet_loss": None, "error": str(e)}

def dns_lookup(domain: str) -> list:
    """Return a list of IP addresses for the domain."""
    try:
        return list(set(socket.gethostbyname_ex(domain)[2]))
    except Exception:
        return []

def traceroute(target: str) -> str:
    """Run tracert and return the raw output as a string."""
    try:
        result = subprocess.run(["tracert", target], capture_output=True, text=True, timeout=30)
        return result.stdout.strip()
    except Exception as e:
        return str(e)

def check_ports(target: str, ports: list) -> list:
    """Check if specified ports are open on the target. Returns list of open ports."""
    open_ports = []
    for port in ports:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            s.settimeout(2)
            try:
                if s.connect_ex((target, port)) == 0:
                    open_ports.append(port)
            except Exception:
                continue
    return open_ports

def run_network_check(target: str, ports: list = None) -> str:
    """High‑level helper that runs all checks and logs them. Returns a summary string."""
    if ports is None:
        ports = [22, 80, 443]
    data = {"target": target}
    ping = ping_target(target)
    data.update(ping)
    data["dns"] = dns_lookup(target)
    data["tracert"] = traceroute(target)
    data["open_ports"] = check_ports(target, ports)
    _log_to_excel(data)
    return (
        f"Network check for {target}: ping {ping.get('ping_ms')} ms, "
        f"loss {ping.get('packet_loss')}%, open ports {data['open_ports']}."
    )
